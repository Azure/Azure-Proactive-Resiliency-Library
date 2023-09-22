import json
import os
import fnmatch
import re
import argparse
import pandas as pd

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

# Import Azure Resource Graph library
import azure.mgmt.resourcegraph as arg

# Import specific methods and models from other libraries
from azure.identity import AzureCliCredential

# Concurrency for running queries
_max_workers = 5

# Parse command line arguments
def parseArguments():
  parser = argparse.ArgumentParser(description='This tool runs all of the Azure Proactive Resiliency Library queries based on the Azure resource types present in the given set of subscriptions and outputs a populated Excel workbook based on the Action Plan Template.')
  parser.add_argument('-s', '--ids', type=str, metavar='ID', nargs='+', required=True, help='List of Subscription IDs')
  parser.add_argument('-p', '--path', type=str, required=False, help='Path to APRL queries', default='../../docs/content/services')
  parser.add_argument('-j', '--output-json', action="store_true", required=False, help='Output to JSON file')
  parser.add_argument('-c', '--output-csv', action="store_true", required=False, help='Output to CSV file')
  parser.add_argument('-l', '--load-json', type=str, metavar='JSON_FIFLE', required=False, help='Load an JSON file with query results')
  parser.add_argument('-w', '--workers', type=int, required=False, help='Number of concurrent workers', default='5')

  args = parser.parse_args()

  return args

# Query Azure Resource Graph
def queryAzureResourceGraph(query, subscriptionId):
  # Get your credentials from Azure CLI (development only!) and get your subscription list
  credential = AzureCliCredential()

  # Create Azure Resource Graph client and set options
  argClient = arg.ResourceGraphClient(credential)
  argQueryOptions = arg.models.QueryRequestOptions(result_format="objectArray")

  # Create query
  argQuery = arg.models.QueryRequest(subscriptions=[subscriptionId], query=query, options=argQueryOptions)

  # Run query
  argResults = argClient.resources(argQuery)

  # Show Python object
  return argResults

# Get a list of all resource types in subscription list
def getResourceTypes(subscriptionId):
  query = "Resources | summarize by type"
  results = queryAzureResourceGraph(query=query, subscriptionId=subscriptionId)
  return results.data

# Get a list of all APRL queries
def getAprlQueries(path):
  includes = ['*.kql']
  excludes = []
  queries = {}

  for root, dirs, files in os.walk(path, topdown=True):
    # excludes can be done with fnmatch.filter and complementary set,
    # but it's more annoying to read.
    dirs[:] = [d for d in dirs if d not in excludes]
    for pat in includes:
      for file in fnmatch.filter(files, pat):
        with open(os.path.join(root, file), "r") as f:
          content = f.read()

          for x in re.split('\||and', content):
            if x.lower().strip().startswith("where type ="):
              resourceType = x.split("=")[-1].replace('=', '').replace('~', '').replace('\'','').replace('\"','').strip().lower()

              if resourceType not in queries:
                queries[resourceType] = []

              #text_part = file.replace('.kql', '').split('-')[0].upper()
              #num_part = re.sub(r"\D", '', file).zfill(2)

              query = {
                "id": file.replace('.kql', '').upper(),
                "kql": content
              }

              queries[resourceType].append(query)

              # Only need to find the first resource type
              break;
  return queries

def runAprlQuery(query, subscriptionId):
  print(f"Running APRL query: {query['id']}")

  results = []
  error = ""

  try:
    argResults = queryAzureResourceGraph(query['kql'], subscriptionId)
    results = argResults.data
  except Exception as e:
    error = f"Error: {str(e)}"
    #print(error)

  impactedResources = []
  for result in results:
    if 'id' in result:
      impactedResources.append(result['id'])
    elif 'name' in result:
      impactedResources.append(result['name'])

  impactedResources = sorted(list(set(impactedResources)))

  aprlQueryResult = {
    "id": query['id'],
    "count": len(impactedResources),
    "subscriptionId": subscriptionId,
    "impactedResources": impactedResources,
    "results": results,
    "error": error,
  }

  return aprlQueryResult

# Run APRL queries for each resource type found in subscription
def runAprlQueries(aprlQueryResults, resourceTypes, subscriptionId):
  # Loop through each resource type and run the queries
  for item in resourceTypes:
    resourceType = item['type'].lower()
    if resourceType in aprlQueries:
      with ThreadPoolExecutor(max_workers=_max_workers) as executor:
        futures = [executor.submit(runAprlQuery, query=query, subscriptionId=subscriptionId) for query in aprlQueries[resourceType]]
        for future in futures:
            aprlQueryResult = future.result()
            # Only add the query result if it has results
            if aprlQueryResult['count'] > 0:
              aprlQueryResults['queries'].append(aprlQueryResult)

  # Sort the results by query name with padded numberic part to two digits
  #aprlQueryResults['queries'] = sorted(aprlQueryResults['queries'], key=lambda x: x['id'])

  return aprlQueryResults

# Load query results from a JSON file
def loadJsonFile(filename):
  with open(filename, "r") as f:
    data = json.load(f)
  return data

# Output the query results to a JSON file
def OutputToJsonFile(filename, data):
  # Write the results to a file
  with open(filename, "w") as f:
      json.dump(data, f, indent=2)

# Output the query results to a CSV file
def OutputToCsvFile(filename, data):

  df = pd.DataFrame(data['queries'])

  # convert impactedResources column to a string
  df['impactedResources'] = df['impactedResources'].apply(lambda x: '\n'.join(x)).astype(str)

  # convert error column to a string
  df['results'] = df['results'].astype(str)

  # convert error column to a string
  df['error'] = df['error'].astype(str)

  # Save the DataFrame to a CSV file
  df.to_csv(filename, index=False, header=True)

def updateRecommendationsInPlan(ws, idColumnName, aprlQueriesResults, table, startCell):

  # Get the column index of the Recommendation ID column
  idColumnId = next((cell.col_idx for cell in ws[2] if cell.value.lower().strip() == idColumnName.lower().strip()), 0)
  if idColumnId == 0:
    # If the column is not found, raise an exception
    raise Exception(f"Column '{idColumnName}' not found in worksheet '{ws.title}'")

  # Get a list of ids and from the aprlQueriesResults
  ids = [query['id'] for query in aprlQueriesResults['queries']]

  # Delete rows that are not in the list of ids
  for i in range(ws.max_row, 2, -1):
    recommendationId = ws.cell(row=i, column=idColumnId).value.upper()
    if recommendationId not in ids:
      ws.delete_rows(i,1)

  # Loop through remaining rows and set height to 60
  for i in range(2, ws.max_row+1):
    ws.row_dimensions[i].height = 60

  ws.tables[table].ref = f"{startCell}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

  return ws

def addImpactedResourcesToPlan(aprlQueriesResults, ws, table, startCell):
  ws.delete_rows(2, ws.max_row-1)

  for query in aprlQueriesResults['queries']:
    for resource in query['impactedResources']:
      resourceName = resource.split('/')[-1]
      ws.append([query['id'], resourceName, resource])

  ws.tables[table].ref = f"{startCell}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

### Main ###

# Parse the command line arguments
args = parseArguments()

subscriptions = args.ids
pathToServices = args.path
_max_workers = args.workers

now = datetime.now()
dt_string = now.strftime("%m/%d/%Y %H:%M:%S")

aprlQueryResults = {
  "name": "Azure Proactive Resiliency Library Queries",
  "date": dt_string,
  "queries": []
}

# Load the JSON file if specified on the command line
if args.load_json:
  aprlQueryResults = loadJsonFile(args.load_json)
else:
  # Get a list of all APRL queries
  aprlQueries = getAprlQueries(pathToServices)

  # Loop through each subscription and run the queries
  for subscriptionId in subscriptions:
    # Get a list of all resource types in subscription list
    resourceTypes = getResourceTypes(subscriptionId)

    # Run APRL queries for each resource type found in subscription
    aprlQueryResults = runAprlQueries(aprlQueryResults, resourceTypes, subscriptionId)


# Load the template with the complete list of APRL recommendations
templatePath = 'Action Plan - Template.xlsx'
wbPlan = load_workbook(templatePath)
wsRecommendations = wbPlan['Recommendations']
wsImpactedResources = wbPlan['ImpactedResources']

# Filter the APRL recommendations based on the results of the queries
wsRecommendations = updateRecommendationsInPlan(wsRecommendations, "ID", aprlQueryResults, 'Table1', 'A2')

# Add impacted resources to the Action Plan
wsImpactedResources=addImpactedResourcesToPlan(aprlQueryResults, wsImpactedResources, 'Table2', 'A1')

# Save the action plan
outputPath = "Action Plan.xlsx"
wbPlan.save(outputPath)

dt_string = now.strftime("%Y%m%d%H%M%S")

# Output the results to file
if args.output_json and not args.load_json:
  OutputToJsonFile(f"aprlQueryResults_{dt_string}.json", aprlQueryResults)

if args.output_csv:
  OutputToCsvFile(f"aprlQueryResults_{dt_string}.csv", aprlQueryResults)

print("Done!")
